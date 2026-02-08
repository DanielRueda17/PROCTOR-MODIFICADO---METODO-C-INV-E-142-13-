"""Generacion de ficha tecnica en Excel para Proctor Modificado Metodo C."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from calculos_proctor import ResultadoPunto
from datos_proctor_laboratorio import ConfiguracionEnsayo, PuntoCompactacion

COLOR_AZUL = "1F4E78"
COLOR_AZUL_CLARO = "D9E1F2"
COLOR_GRIS_CLARO = "F2F2F2"
COLOR_OK = "C6EFCE"
COLOR_ALERTA = "FCE4D6"
COLOR_INFO = "D9EAD3"
COLOR_BLANCO = "FFFFFF"
BORDE_FINO = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _aplicar_titulo(ws, titulo: str, subtitulo: str) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_AZUL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="404040")
    ws["A2"].fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(ws, max_width: int = 42) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ancho = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            texto = "" if cell.value is None else str(cell.value)
            ancho = max(ancho, len(texto))
        ws.column_dimensions[col_letter].width = min(max(ancho + 2, 11), max_width)


def _crear_tabla(
    ws,
    row_inicio: int,
    headers: list[str],
    filas: list[list[object]],
    nombre_tabla: str,
) -> tuple[int, int]:
    row = row_inicio
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill("solid", fgColor=COLOR_AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDE_FINO

    for fila in filas:
        row += 1
        for col, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDE_FINO
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)

    col_fin = len(headers)
    ref = f"A{row_inicio}:{ws.cell(row=row, column=col_fin).coordinate}"
    tabla = Table(displayName=nombre_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)
    return row_inicio, row


def _crear_sheet_ficha(
    wb: Workbook,
    config: ConfiguracionEnsayo,
    volumen_cm3: float,
    diametro_promedio_cm: float,
    altura_promedio_cm: float,
    resultados_validos: list[ResultadoPunto],
    optimo: dict,
    mensajes_validacion: list[str],
) -> None:
    ws = wb.active
    ws.title = "Ficha_Tecnica"
    _aplicar_titulo(
        ws,
        "FICHA TECNICA - PROCTOR MODIFICADO (METODO C)",
        "Norma INV E-142-13 | Generado: " + datetime.now().strftime("%Y-%m-%d %H:%M"),
    )

    ws["A4"] = "Resumen de ensayo"
    ws["A4"].font = Font(bold=True, color=COLOR_AZUL)

    resumen = [
        ("Norma", config.norma),
        ("Metodo", config.metodo),
        ("Preparacion", config.preparacion_muestra),
        ("Capas", config.capas),
        ("Golpes por capa", config.golpes_por_capa),
        ("Volumen de molde (cm3)", round(volumen_cm3, 2)),
        ("Diametro promedio (cm)", round(diametro_promedio_cm, 3)),
        ("Altura promedio (cm)", round(altura_promedio_cm, 3)),
        ("Puntos validos", len(resultados_validos)),
        ("Metodo de estimacion del optimo", optimo["metodo"]),
        ("Humedad optima, w_opt (%)", round(optimo["humedad_optima_pct"], 3)),
        ("Peso unitario seco maximo, gamma_d,max (kN/m3)", round(optimo["peso_unitario_seco_max_kn_m3"], 3)),
    ]

    row = 5
    for etiqueta, valor in resumen:
        ws[f"A{row}"] = etiqueta
        ws[f"B{row}"] = valor
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLOR_AZUL_CLARO)
        ws[f"A{row}"].border = BORDE_FINO
        ws[f"B{row}"].border = BORDE_FINO
        row += 1

    ws[f"A{row+1}"] = "Analisis tecnico"
    ws[f"A{row+1}"].font = Font(bold=True, color=COLOR_AZUL)

    analisis = _generar_analisis(resultados_validos, optimo, mensajes_validacion)
    for i, linea in enumerate(analisis, start=row + 2):
        ws[f"A{i}"] = f"- {linea}"
        ws[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2 + len(analisis), end_column=8)

    _autofit_columns(ws)


def _generar_analisis(
    resultados_validos: list[ResultadoPunto],
    optimo: dict,
    mensajes_validacion: list[str],
) -> list[str]:
    lineas: list[str] = []
    lineas.append(
        f"Se procesaron {len(resultados_validos)} puntos validos para la curva de compactacion."
    )
    lineas.append(
        "El optimo reportado se obtuvo con: "
        f"{optimo['metodo']} (w_opt={optimo['humedad_optima_pct']:.3f} %, "
        f"gamma_d,max={optimo['peso_unitario_seco_max_kn_m3']:.3f} kN/m3)."
    )

    alertas = [m for m in mensajes_validacion if m.startswith("ALERTA")]
    if alertas:
        lineas.append(
            "Se identificaron alertas de calidad/representatividad; revisar antes de usar "
            "el resultado como valor definitivo de diseno."
        )
    else:
        lineas.append(
            "No se detectaron alertas criticas en la validacion basica, por lo que el resultado "
            "puede usarse como referencia tecnica del ensayo."
        )
    return lineas


def _crear_sheet_datos(wb: Workbook, puntos: list[PuntoCompactacion]) -> None:
    ws = wb.create_sheet("Datos_Laboratorio")
    _aplicar_titulo(
        ws,
        "DATOS DE LABORATORIO",
        "Entradas originales usadas en los calculos del ensayo",
    )

    headers = [
        "Punto",
        "w_obj (%)",
        "Masa inicial (g)",
        "Agua adicionada (g)",
        "Wb (g)",
        "Wb+WMh (g)",
        "Wr (g)",
        "Wr+Wmh (g)",
        "Wr+Wms (g)",
    ]
    filas: list[list[object]] = []
    for p in puntos:
        filas.append(
            [
                p.nombre,
                p.humedad_objetivo_pct,
                p.masa_muestra_inicial_g,
                p.agua_adicionada_g,
                None if p.WM is None else p.WM.Wb_g,
                None if p.WM is None else p.WM.Wb_WMh_g,
                None if p.Wm is None else p.Wm.Wr_g,
                None if p.Wm is None else p.Wm.Wr_Wmh_g,
                None if p.Wm is None else p.Wm.Wr_Wms_g,
            ]
        )

    _crear_tabla(ws, row_inicio=4, headers=headers, filas=filas, nombre_tabla="TablaDatosLab")
    _autofit_columns(ws)


def _crear_sheet_calculos(wb: Workbook, resultados: list[ResultadoPunto]) -> None:
    ws = wb.create_sheet("Calculos")
    _aplicar_titulo(
        ws,
        "RESULTADOS DE CALCULO",
        "Valores calculados por punto de compactacion",
    )

    headers = [
        "Punto",
        "w_obj (%)",
        "w_usada (%)",
        "Origen de w",
        "WMh (g)",
        "rho_h (g/cm3)",
        "rho_d (g/cm3)",
        "gamma_h (kN/m3)",
        "gamma_d (kN/m3)",
        "Estado",
    ]
    filas: list[list[object]] = []
    for r in resultados:
        filas.append(
            [
                r.nombre,
                r.humedad_objetivo_pct,
                r.humedad_usada_pct,
                r.origen_humedad,
                r.masa_suelo_humedo_molde_g,
                r.densidad_humeda_g_cm3,
                r.densidad_seca_g_cm3,
                r.peso_unitario_humedo_kn_m3,
                r.peso_unitario_seco_kn_m3,
                r.estado,
            ]
        )

    inicio, fin = _crear_tabla(ws, row_inicio=4, headers=headers, filas=filas, nombre_tabla="TablaCalculos")

    # Colorear estado para lectura rapida.
    col_estado = 10
    for row in range(inicio + 1, fin + 1):
        celda = ws.cell(row=row, column=col_estado)
        texto = "" if celda.value is None else str(celda.value)
        if texto.startswith("OK"):
            celda.fill = PatternFill("solid", fgColor=COLOR_OK)
        elif texto.startswith("ALERTA") or texto.startswith("ERROR"):
            celda.fill = PatternFill("solid", fgColor=COLOR_ALERTA)
        else:
            celda.fill = PatternFill("solid", fgColor=COLOR_INFO)

    _autofit_columns(ws)


def _crear_sheet_validacion(wb: Workbook, mensajes_validacion: list[str]) -> None:
    ws = wb.create_sheet("Validacion")
    _aplicar_titulo(
        ws,
        "VALIDACION NORMATIVA",
        "Chequeos basicos de conformidad INV E-142-13",
    )

    headers = ["Tipo", "Mensaje"]
    filas: list[list[object]] = []
    for msg in mensajes_validacion:
        if msg.startswith("OK"):
            tipo = "OK"
        elif msg.startswith("ALERTA"):
            tipo = "ALERTA"
        else:
            tipo = "INFO"
        filas.append([tipo, msg])

    inicio, fin = _crear_tabla(
        ws,
        row_inicio=4,
        headers=headers,
        filas=filas,
        nombre_tabla="TablaValidacion",
    )

    for row in range(inicio + 1, fin + 1):
        tipo = ws.cell(row=row, column=1).value
        if tipo == "OK":
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COLOR_OK)
        elif tipo == "ALERTA":
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COLOR_ALERTA)
        else:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COLOR_INFO)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["B"].width = 90
    _autofit_columns(ws)


def _crear_sheet_curva(wb: Workbook, resultados_validos: list[ResultadoPunto], optimo: dict) -> None:
    ws = wb.create_sheet("Curva_Compactacion")
    _aplicar_titulo(
        ws,
        "CURVA DE COMPACTACION",
        "Relacion humedad de moldeo vs peso unitario seco",
    )

    ws["A4"] = "w (%)"
    ws["B4"] = "gamma_d (kN/m3)"
    ws["D4"] = "w_opt (%)"
    ws["E4"] = "gamma_d,max (kN/m3)"
    ws["G4"] = "w_ajuste (%)"
    ws["H4"] = "gamma_d_ajuste (kN/m3)"
    for cell in ("A4", "B4", "D4", "E4", "G4", "H4"):
        ws[cell].font = Font(bold=True, color=COLOR_BLANCO)
        ws[cell].fill = PatternFill("solid", fgColor=COLOR_AZUL)
        ws[cell].alignment = Alignment(horizontal="center")

    data = sorted(
        [(r.humedad_usada_pct, r.peso_unitario_seco_kn_m3) for r in resultados_validos],
        key=lambda x: x[0],
    )
    row = 5
    for w, g in data:
        ws.cell(row=row, column=1, value=float(w))
        ws.cell(row=row, column=2, value=float(g))
        row += 1
    row_fin_puntos = row - 1

    ws["D5"] = float(optimo["humedad_optima_pct"])
    ws["E5"] = float(optimo["peso_unitario_seco_max_kn_m3"])

    row_fin_ajuste = 4
    if len(data) >= 3:
        a, b, c = np.polyfit(
            np.array([x for x, _ in data], dtype=float),
            np.array([y for _, y in data], dtype=float),
            2,
        )
        w_min = min(x[0] for x in data)
        w_max = max(x[0] for x in data)
        w_plot = np.linspace(w_min, w_max, 100)
        g_plot = a * w_plot**2 + b * w_plot + c
        row_aj = 5
        for w, g in zip(w_plot, g_plot):
            ws.cell(row=row_aj, column=7, value=float(w))
            ws.cell(row=row_aj, column=8, value=float(g))
            row_aj += 1
        row_fin_ajuste = row_aj - 1

    chart = ScatterChart()
    chart.scatterStyle = "lineMarker"
    chart.title = "Curva de compactacion Proctor Modificado - Metodo C"
    chart.x_axis.title = "Humedad de moldeo, w (%)"
    chart.y_axis.title = "Peso unitario seco, gamma_d (kN/m3)"
    chart.height = 11
    chart.width = 18
    chart.legend.position = "r"

    x_ref = Reference(ws, min_col=1, min_row=5, max_row=row_fin_puntos)
    y_ref = Reference(ws, min_col=2, min_row=5, max_row=row_fin_puntos)
    serie_puntos = Series(y_ref, x_ref, title="Puntos medidos")
    serie_puntos.marker.symbol = "circle"
    serie_puntos.marker.size = 7
    serie_puntos.graphicalProperties.line.noFill = True
    chart.series.append(serie_puntos)

    serie_suave = Series(y_ref, x_ref, title="Linea suavizada")
    serie_suave.marker.symbol = "none"
    serie_suave.smooth = True
    serie_suave.graphicalProperties.line.solidFill = "000000"
    serie_suave.graphicalProperties.line.prstDash = "sysDot"
    serie_suave.graphicalProperties.line.width = 24000
    chart.series.append(serie_suave)

    if row_fin_ajuste > 4:
        x_aj = Reference(ws, min_col=7, min_row=5, max_row=row_fin_ajuste)
        y_aj = Reference(ws, min_col=8, min_row=5, max_row=row_fin_ajuste)
        serie_ajuste = Series(y_aj, x_aj, title="Regresion parabolica")
        serie_ajuste.marker.symbol = "none"
        serie_ajuste.smooth = True
        serie_ajuste.graphicalProperties.line.solidFill = "C00000"
        serie_ajuste.graphicalProperties.line.prstDash = "sysDot"
        serie_ajuste.graphicalProperties.line.width = 24000
        chart.series.append(serie_ajuste)

    x_opt = Reference(ws, min_col=4, min_row=5, max_row=5)
    y_opt = Reference(ws, min_col=5, min_row=5, max_row=5)
    serie_optimo = Series(y_opt, x_opt, title="Punto optimo")
    serie_optimo.marker.symbol = "diamond"
    serie_optimo.marker.size = 9
    serie_optimo.graphicalProperties.line.solidFill = "000000"
    chart.series.append(serie_optimo)

    # Escalas del eje X (humedad) para lectura tecnica.
    x_vals = [x for x, _ in data] + [float(optimo["humedad_optima_pct"])]
    x_min = min(x_vals)
    x_max = max(x_vals)
    x_span = max(x_max - x_min, 0.5)
    x_major = 0.2 if x_span <= 3 else 0.5 if x_span <= 6 else 1.0
    x_pad = max(0.2, 0.08 * x_span)
    chart.x_axis.scaling.min = round(x_min - x_pad, 3)
    chart.x_axis.scaling.max = round(x_max + x_pad, 3)
    chart.x_axis.majorUnit = x_major
    chart.x_axis.minorUnit = x_major / 2

    # Escalas del eje Y (gamma_d) para lectura tecnica.
    y_vals = [y for _, y in data] + [float(optimo["peso_unitario_seco_max_kn_m3"])]
    y_min = min(y_vals)
    y_max = max(y_vals)
    y_span = max(y_max - y_min, 0.2)
    y_major = 0.05 if y_span <= 0.6 else 0.1 if y_span <= 1.5 else 0.2
    y_pad = max(0.05, 0.10 * y_span)
    chart.y_axis.scaling.min = round(y_min - y_pad, 3)
    chart.y_axis.scaling.max = round(y_max + y_pad, 3)
    chart.y_axis.majorUnit = y_major
    chart.y_axis.minorUnit = y_major / 2

    ws.add_chart(chart, "A8")
    ws["A6"] = ws["A6"].value  # mantiene hoja con contenido visible en algunos visores.
    _autofit_columns(ws)


def _crear_sheet_referencias(wb: Workbook, referencias: Iterable[str]) -> None:
    ws = wb.create_sheet("Referencias")
    _aplicar_titulo(
        ws,
        "REFERENCIAS Y TRAZABILIDAD",
        "Base normativa y criterios aplicados en el informe",
    )
    ws["A4"] = "Listado de referencias"
    ws["A4"].font = Font(bold=True, color=COLOR_AZUL)

    row = 6
    for linea in referencias:
        if not linea.strip():
            row += 1
            continue
        ws[f"A{row}"] = linea
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 120


def exportar_ficha_tecnica_excel(
    ruta_excel: Path,
    config: ConfiguracionEnsayo,
    molde: dict,
    puntos: list[PuntoCompactacion],
    resultados: list[ResultadoPunto],
    resultados_validos: list[ResultadoPunto],
    optimo: dict,
    mensajes_validacion: list[str],
    texto_referencias: str,
) -> None:
    """Crea un Excel tecnico con tablas, estilos y grafica de compactacion."""

    volumen_cm3 = float(molde["volumen_cm3"])
    diametro_promedio_cm = float(molde["diametro_promedio_cm"])
    altura_promedio_cm = float(molde["altura_promedio_cm"])

    wb = Workbook()
    _crear_sheet_ficha(
        wb=wb,
        config=config,
        volumen_cm3=volumen_cm3,
        diametro_promedio_cm=diametro_promedio_cm,
        altura_promedio_cm=altura_promedio_cm,
        resultados_validos=resultados_validos,
        optimo=optimo,
        mensajes_validacion=mensajes_validacion,
    )
    _crear_sheet_datos(wb, puntos)
    _crear_sheet_calculos(wb, resultados)
    _crear_sheet_validacion(wb, mensajes_validacion)
    _crear_sheet_curva(wb, resultados_validos, optimo)
    _crear_sheet_referencias(wb, texto_referencias.splitlines())

    ruta_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_excel)
